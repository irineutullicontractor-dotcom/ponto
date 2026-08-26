import io
import re
from copy import copy

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURAÇÃO
# ============================================================

st.set_page_config(
    page_title="Automação Ponto S. Diogo",
    page_icon="🕘",
    layout="wide",
)

st.title("🕘 Automação do Ponto — S. Diogo")
st.caption(
    "Preenche automaticamente a planilha PONTO S.DIOGO a partir do Cartão Ponto."
)


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

def normalizar_texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def normalizar_comparacao(valor):
    texto = normalizar_texto(valor)
    texto = re.sub(r"\s+", " ", texto)
    return texto.casefold()


def numero_limpo(valor):
    if valor is None:
        return ""
    texto = str(valor).strip()
    if texto.endswith(".0"):
        texto = texto[:-2]
    return re.sub(r"\D", "", texto)


def converter_horas(valor):
    """
    Mantém os valores de horas da origem em formato adequado
    para serem gravados no Excel.

    Aceita:
    - datetime/time
    - valores numéricos
    - textos como 08:30
    - textos vazios
    """
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None

    if hasattr(valor, "hour") and hasattr(valor, "minute"):
        return valor

    if isinstance(valor, (int, float)):
        if pd.isna(valor):
            return None
        # Excel normalmente representa duração como fração de dia.
        return float(valor)

    texto = str(valor).strip()
    if not texto or texto.lower() in {"nan", "none", "-"}:
        return None

    # Aceita H:MM ou HH:MM
    match = re.fullmatch(r"(\d{1,3}):(\d{2})(?::(\d{2}))?", texto)
    if match:
        h = int(match.group(1))
        m = int(match.group(2))
        s = int(match.group(3) or 0)
        return (h * 3600 + m * 60 + s) / 86400

    # Retorna o valor original se não for um horário reconhecido.
    return valor


def copiar_estilo(celula_origem, celula_destino):
    """
    Copia somente propriedades visuais, preservando fórmulas/valores
    da célula de destino quando necessário.
    """
    if celula_origem.has_style:
        celula_destino._style = copy(celula_origem._style)
    if celula_origem.number_format:
        celula_destino.number_format = celula_origem.number_format
    if celula_origem.alignment:
        celula_destino.alignment = copy(celula_origem.alignment)
    if celula_origem.protection:
        celula_destino.protection = copy(celula_origem.protection)


# ============================================================
# LEITURA DA ORIGEM
# ============================================================

def encontrar_colaborador_e_matricula(ws, linha):
    """
    Na estrutura do Cartão Ponto:
      A = nome do colaborador
      M = matrícula

    A função tenta encontrar nome e matrícula na região inicial
    de cada bloco de colaborador.
    """
    nome = normalizar_texto(ws.cell(linha, 1).value)
    matricula = numero_limpo(ws.cell(linha, 13).value)

    if nome and nome.casefold() != "contractor engenharia ltda":
        return nome, matricula

    # Procura nas próximas linhas do bloco.
    for r in range(linha + 1, min(linha + 12, ws.max_row + 1)):
        valor_a = normalizar_texto(ws.cell(r, 1).value)
        valor_m = numero_limpo(ws.cell(r, 13).value)

        if valor_a and valor_a.casefold() != "contractor engenharia ltda":
            # Evita pegar a linha de datas.
            if not re.search(r"\d{1,2}/\d{1,2}/\d{4}", valor_a):
                return valor_a, valor_m

    return "", ""


def eh_linha_data(valor):
    if valor is None:
        return False

    texto = normalizar_texto(valor)
    return bool(re.search(r"\b\d{1,2}/\d{1,2}/\d{4}\b", texto))


def extrair_dia(valor):
    if valor is None:
        return None

    if hasattr(valor, "day"):
        return int(valor.day)

    texto = normalizar_texto(valor)
    match = re.search(r"\b(\d{1,2})/\d{1,2}/\d{4}\b", texto)
    if match:
        return int(match.group(1))

    return None


def identificar_status_dia(ws_origem, linha):
    """
    REGRA DE FALTA/ATESTADO

    Antes de consultar as horas extras, compara:
      E x P

    Para ser válido, E e P precisam ser iguais.

    Se forem iguais:
      - contendo 'FALTA'   -> F
      - contendo 'Atestad' -> AT

    Se E e P forem diferentes, não lança F/AT.
    """
    valor_e = normalizar_comparacao(ws_origem.cell(linha, 5).value)   # E
    valor_p = normalizar_comparacao(ws_origem.cell(linha, 16).value)  # P

    if not valor_e or not valor_p or valor_e != valor_p:
        return ""

    if "falta" in valor_e:
        return "F"

    if "atestad" in valor_e:
        return "AT"

    return ""


def extrair_colaboradores(ws):
    """
    Percorre todos os blocos iniciados por CONTRACTOR ENGENHARIA LTDA.
    """
    colaboradores = []

    marcadores = []
    for linha in range(1, ws.max_row + 1):
        valor = normalizar_texto(ws.cell(linha, 1).value)
        if valor.casefold() == "contractor engenharia ltda":
            marcadores.append(linha)

    for i, inicio in enumerate(marcadores):
        fim = marcadores[i + 1] - 1 if i + 1 < len(marcadores) else ws.max_row

        nome = ""
        matricula = ""
        linha_nome = None

        # Procura nome/matrícula dentro do bloco.
        for r in range(inicio + 1, min(inicio + 15, fim + 1)):
            a = normalizar_texto(ws.cell(r, 1).value)
            m = numero_limpo(ws.cell(r, 13).value)

            if a and not eh_linha_data(a):
                if a.casefold() != "contractor engenharia ltda":
                    # Evita cabeçalhos como TOTAIS.
                    if a.casefold() != "totais":
                        nome = a
                        matricula = m
                        linha_nome = r
                        break

        if not nome:
            continue

        dias = {}
        totais = {
            "he50": None,
            "he70": None,
            "he120": None,
            "atrasos": None,
            "faltas": None,
        }

        for r in range(inicio, fim + 1):
            valor_a = normalizar_texto(ws.cell(r, 1).value)

            # Linha diária
            if eh_linha_data(valor_a):
                dia = extrair_dia(valor_a)
                if dia is None:
                    continue

                status = identificar_status_dia(ws, r)

                # Se houver F/AT válido, ele tem prioridade.
                if status:
                    dias[dia] = {
                        "status": status,
                        "he50": None,
                        "he70": None,
                        "he120": None,
                    }
                else:
                    dias[dia] = {
                        "status": "",
                        "he50": converter_horas(ws.cell(r, 24).value),   # X
                        "he70": converter_horas(ws.cell(r, 27).value),   # AA
                        "he120": converter_horas(ws.cell(r, 31).value),  # AE
                    }

            # Linha TOTAIS
            if valor_a.casefold() == "totais":
                totais = {
                    "he50": converter_horas(ws.cell(r, 24).value),   # X
                    "he70": converter_horas(ws.cell(r, 27).value),   # AA
                    "he120": converter_horas(ws.cell(r, 31).value),  # AE
                    "atrasos": converter_horas(ws.cell(r, 34).value), # AH
                    "faltas": converter_horas(ws.cell(r, 36).value),  # AJ
                }

        colaboradores.append(
            {
                "nome": nome,
                "matricula": matricula,
                "dias": dias,
                "totais": totais,
                "linha_nome": linha_nome,
            }
        )

    return colaboradores


# ============================================================
# LEITURA DO MODELO DE DESTINO
# ============================================================

def identificar_colunas_dias(ws_destino):
    """
    Na planilha modelo:
      H em diante = dias do período.
    Lê o número do dia e cria:
      {dia: coluna_excel}
    """
    mapa = {}

    for col in range(8, ws_destino.max_column + 1):
        valor = ws_destino.cell(2, col).value

        if valor is None:
            continue

        dia = None

        if hasattr(valor, "day"):
            dia = int(valor.day)
        else:
            texto = normalizar_texto(valor)
            match = re.search(r"\b(\d{1,2})\b", texto)
            if match:
                candidato = int(match.group(1))
                if 1 <= candidato <= 31:
                    dia = candidato

        if dia is not None:
            mapa[dia] = col

    return mapa


def identificar_linhas_colaboradores(ws_destino):
    """
    Retorna os colaboradores encontrados no modelo.
    A = nome
    B = matrícula
    """
    registros = []

    for linha in range(3, ws_destino.max_row + 1):
        nome = normalizar_texto(ws_destino.cell(linha, 1).value)
        matricula = numero_limpo(ws_destino.cell(linha, 2).value)

        if nome or matricula:
            registros.append(
                {
                    "linha": linha,
                    "nome": nome,
                    "matricula": matricula,
                }
            )

    return registros


# ============================================================
# CRUZAMENTO
# ============================================================

def cruzar_colaborador(origem, destino):
    """
    Prioridade:
      1. matrícula
      2. nome exato normalizado

    Retorna:
      (linha_destino, método, confiança)
    """
    matricula_origem = origem["matricula"]
    nome_origem = normalizar_comparacao(origem["nome"])

    # 1. Matrícula
    if matricula_origem:
        candidatos = [
            d for d in destino
            if d["matricula"] and d["matricula"] == matricula_origem
        ]
        if len(candidatos) == 1:
            return candidatos[0]["linha"], "matrícula", "alta"

    # 2. Nome
    candidatos = [
        d for d in destino
        if d["nome"] and normalizar_comparacao(d["nome"]) == nome_origem
    ]

    if len(candidatos) == 1:
        return candidatos[0]["linha"], "nome", "média"

    return None, "", "não encontrado"


# ============================================================
# PREENCHIMENTO
# ============================================================

def preencher_planilha(origem_bytes, destino_bytes):
    wb_origem = load_workbook(io.BytesIO(origem_bytes), data_only=True)
    ws_origem = wb_origem.active

    wb_destino = load_workbook(io.BytesIO(destino_bytes))
    ws_destino = wb_destino.active

    colaboradores = extrair_colaboradores(ws_origem)
    destino = identificar_linhas_colaboradores(ws_destino)
    mapa_dias = identificar_colunas_dias(ws_destino)

    resultado = []
    contadores = {
        "colaboradores_origem": len(colaboradores),
        "encontrados": 0,
        "nao_encontrados": 0,
        "por_matricula": 0,
        "por_nome": 0,
        "F": 0,
        "AT": 0,
        "he50_dias": 0,
        "he70_dias": 0,
        "he120_dias": 0,
    }

    for origem in colaboradores:
        linha_destino, metodo, confianca = cruzar_colaborador(
            origem, destino
        )

        if linha_destino is None:
            contadores["nao_encontrados"] += 1
            resultado.append(
                {
                    "Nome origem": origem["nome"],
                    "Matrícula origem": origem["matricula"],
                    "Status": "NÃO ENCONTRADO",
                    "Método": "",
                }
            )
            continue

        contadores["encontrados"] += 1
        if metodo == "matrícula":
            contadores["por_matricula"] += 1
        elif metodo == "nome":
            contadores["por_nome"] += 1

        # Totais
        totais = origem["totais"]

        # D = faltas
        if totais["faltas"] is not None:
            ws_destino.cell(linha_destino, 4).value = totais["faltas"]

        # E/F/G são normalmente fórmulas do modelo.
        # Não sobrescrevemos as fórmulas existentes.
        #
        # Se a célula estiver vazia, gravamos o total da origem.
        for col_destino, chave in [
            (5, "he50"),
            (6, "he70"),
            (7, "he120"),
        ]:
            cel = ws_destino.cell(linha_destino, col_destino)
            if cel.value is None and totais[chave] is not None:
                cel.value = totais[chave]

        # AM = atrasos
        if totais["atrasos"] is not None:
            ws_destino.cell(linha_destino, 39).value = totais["atrasos"]

        # Dias
        for dia, dados in origem["dias"].items():
            if dia not in mapa_dias:
                continue

            col = mapa_dias[dia]
            cel = ws_destino.cell(linha_destino, col)

            # F / AT têm prioridade sobre horas extras.
            if dados["status"] == "F":
                cel.value = "F"
                contadores["F"] += 1

            elif dados["status"] == "AT":
                cel.value = "AT"
                contadores["AT"] += 1

            else:
                # Somente escreve horas quando houver valor.
                # Não apaga informações já existentes no modelo.
                if dados["he50"] is not None:
                    cel.value = dados["he50"]
                    contadores["he50_dias"] += 1

                elif dados["he70"] is not None:
                    cel.value = dados["he70"]
                    contadores["he70_dias"] += 1

                elif dados["he120"] is not None:
                    cel.value = dados["he120"]
                    contadores["he120_dias"] += 1

        resultado.append(
            {
                "Nome origem": origem["nome"],
                "Matrícula origem": origem["matricula"],
                "Status": "OK",
                "Método": metodo,
            }
        )

    output = io.BytesIO()
    wb_destino.save(output)
    output.seek(0)

    return output.getvalue(), contadores, pd.DataFrame(resultado)


# ============================================================
# INTERFACE
# ============================================================

st.markdown(
    """
### Como utilizar

1. Envie o **Cartão Ponto fechamento** exportado do sistema.
2. Envie o **PONTO S.DIOGO** em branco que será utilizado como modelo.
3. Clique em **Processar ponto**.
4. Confira o resumo e eventuais colaboradores não encontrados.
5. Baixe a planilha preenchida.
"""
)

col1, col2 = st.columns(2)

with col1:
    arquivo_origem = st.file_uploader(
        "1️⃣ Cartão Ponto — arquivo de origem",
        type=["xlsx"],
        key="origem",
    )

with col2:
    arquivo_destino = st.file_uploader(
        "2️⃣ PONTO S.DIOGO — modelo em branco",
        type=["xlsx"],
        key="destino",
    )

if st.button("🚀 Processar ponto", type="primary"):
    if not arquivo_origem or not arquivo_destino:
        st.error("Envie os dois arquivos antes de processar.")
        st.stop()

    try:
        with st.spinner("Processando cartão ponto..."):
            arquivo_final, contadores, df_resultado = preencher_planilha(
                arquivo_origem.getvalue(),
                arquivo_destino.getvalue(),
            )

        st.success("✅ Planilha processada com sucesso!")

        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Colaboradores origem", contadores["colaboradores_origem"])
        c2.metric("Encontrados", contadores["encontrados"])
        c3.metric("Não encontrados", contadores["nao_encontrados"])
        c4.metric("F / AT", contadores["F"] + contadores["AT"])

        st.markdown("### 📊 Lançamentos realizados")

        a, b, c, d = st.columns(4)
        a.metric("Faltas — F", contadores["F"])
        b.metric("Atestados — AT", contadores["AT"])
        c.metric("HE 50% nos dias", contadores["he50_dias"])
        d.metric("HE 70% / 120%", contadores["he70_dias"] + contadores["he120_dias"])

        problemas = df_resultado[df_resultado["Status"] != "OK"]

        if not problemas.empty:
            st.warning(
                f"⚠️ {len(problemas)} colaborador(es) não foram encontrados "
                "no modelo de destino."
            )
            st.dataframe(problemas, use_container_width=True)

        st.markdown("### 🔎 Conferência dos colaboradores")
        st.dataframe(df_resultado, use_container_width=True)

        st.download_button(
            "📥 Baixar PONTO S.DIOGO preenchido",
            data=arquivo_final,
            file_name="PONTO_S_DIOGO_PREENCHIDO.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
        )

    except Exception as e:
        st.error(f"❌ Erro durante o processamento: {e}")
        st.exception(e)
